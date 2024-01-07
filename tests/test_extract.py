import unittest
from pathlib import Path

from openpyxl import Workbook

from playground.ppt_extract import read_ppt


class MyTestCase(unittest.TestCase):
    def test_something(self):
        import_dir = Path().cwd() / '..'
        export_dir = Path().cwd() / '..'

        wb = Workbook()
        ws = Workbook().create_sheet("test")

        for i in import_dir.iterdir():
            if i.is_file() and i.suffix == ".pptx":
                read_ppt(i, ws)

        wb.save(export_dir / 'testexcel.xlsx')

        textexcel = export_dir / 'textexcel.xlsx'

        self.assertIsNotNone(textexcel)
        self.addCleanup(remove_file)


def remove_file():
    saved_file = Path().cwd() / '..' / 'testexcel.xlsx'
    saved_file.unlink()


if __name__ == '__main__':
    unittest.main()
